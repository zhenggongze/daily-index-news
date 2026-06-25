"""
从15个白名单信源采集AI算力产业链新闻
输出：ai_news_for_labeling.xlsx
"""

import feedparser
import time
import random
import sys
import os
import re
import requests
import warnings
import urllib3
from datetime import datetime, timedelta

warnings.filterwarnings("ignore", category=urllib3.exceptions.InsecureRequestWarning)

BASE = os.path.dirname(os.path.abspath(__file__))
SKILLS_DIR = os.path.join(BASE, ".trae", "skills", "a-stock-data")
sys.path.insert(0, SKILLS_DIR)

from scripts.news import global_news as eastmoney_news

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    os.system("pip install openpyxl -q")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ========== 配置区 ==========
FETCH_DAYS = 30  # 回溯天数

# RSS信源配置
RSS_SOURCES = {
    "华尔街见闻": {
        "url": "https://feed.wallstreetcn.com/news/global",
        "mainlines": ["B", "D"],
        "priority": 1,
        "alt_urls": []
    },
    "澎湃新闻": {
        "url": "https://www.thepaper.cn/rss/news.xml",
        "mainlines": ["A"],
        "priority": 2,
        "alt_urls": ["https://www.thepaper.cn/rss/feed.xml"]
    },
    "界面新闻": {
        "url": "https://www.jiemian.com/rss.xml",
        "mainlines": ["C", "D"],
        "priority": 3,
        "alt_urls": ["https://a.jiemian.com/index.php?m=content&c=index&a=jsonp_rss"]
    },
    "IT之家": {
        "url": "https://www.ithome.com/rss.xml",
        "mainlines": ["C", "D"],
        "priority": 3,
        "alt_urls": ["https://feed.ithome.com/", "https://www.ithome.com/feed/"]
    },
    "爱范儿": {
        "url": "https://www.ifanr.com/feed",
        "mainlines": ["D"],
        "priority": 4,
        "alt_urls": []
    },
    "199IT": {
        "url": "https://www.199it.com/feed",
        "mainlines": ["D"],
        "priority": 4,
        "alt_urls": []
    },
    "新华网": {
        "url": "http://www.xinhuanet.com/rss/news.xml",
        "mainlines": ["A"],
        "priority": 2,
        "alt_urls": ["http://www.news.cn/rss/news.xml"]
    },
    "SemiAnalysis": {
        "url": "https://semianalysis.substack.com/feed",
        "mainlines": ["B"],
        "priority": 1,
        "alt_urls": []
    },
    "NVIDIA Blog": {
        "url": "https://developer.nvidia.com/blog/feed",
        "mainlines": ["B"],
        "priority": 1,
        "alt_urls": []
    },
    "Tom's Hardware": {
        "url": "https://www.tomshardware.com/feeds/all",
        "mainlines": ["B"],
        "priority": 3,
        "alt_urls": []
    },
    "ServeTheHome": {
        "url": "https://www.servethehome.com/feed",
        "mainlines": ["B"],
        "priority": 3,
        "alt_urls": []
    },
    "新浪财经": {
        "url": "https://feed.mix.sina.com.cn/api/roll/get?pageid=153&lid=2516&num=20",
        "mainlines": ["A", "B", "D"],
        "priority": 2,
        "alt_urls": []
    }
}

# AI算力产业链 关键词体系
KEYWORDS_MAINLINE = {
    "A": {
        "name": "国产替代自主可控",
        "keywords": [
            "国产替代", "自主可控", "半导体", "芯片", "晶圆", "光刻", "大基金",
            "信创", "EDA", "IP核", "HBM", "先进封装", "存储芯片", "设备国产",
            "中芯国际", "华大九天", "北方华创", "中微公司", "长鑫存储", "长江存储",
            "龙芯", "飞腾", "海光", "寒武纪", "昇腾", "华为", "麒麟",
            "国产GPU", "国产AI芯片", "国产化率", "渗透率", "良率突破",
            "光刻机", "刻蚀机", "薄膜沉积", "清洗设备", "涂胶显影",
            "Chiplet", "2.5D封装", "3D封装", "TSV", "硅通孔",
            "国芯", "景嘉微", "澜起科技", "兆易创新", "通富微电"
        ]
    },
    "B": {
        "name": "英伟达全球供应链",
        "keywords": [
            "NVIDIA", "英伟达", "GB200", "B300", "B200", "NVL72", "NVL36",
            "Rubin", "Blackwell", "Hopper",
            "光模块", "光互联", "CPO", "硅光", "LPO",
            "PCB", "覆铜板", "高速铜缆", "DAC", "AEC",
            "散热", "液冷", "浸没式液冷", "冷板式液冷",
            "服务器", "GPU服务器", "AI服务器",
            "H100", "H200", "B100", "A100",
            "CoWoS", "先进封装产能",
            "一供", "二供", "份额提升", "ASP提升",
            "中际旭创", "天孚通信", "新易盛", "工业富联", "沪电股份",
            "胜宏科技", "深南电路", "生益科技", "立讯精密",
            "英伟达认证", "NVIDIA认证", "数据中心GPU",
            "800G", "1.6T", "1.6T光模块",
            "Networking", "NVLink", "InfiniBand", "Spectrum-X"
        ]
    },
    "C": {
        "name": "具身智能产业化",
        "keywords": [
            "人形机器人", "具身智能", "灵巧手", "力矩传感器", "六维力传感器",
            "减速器", "谐波减速器", "RV减速器", "丝杠", "滚珠丝杠", "行星滚柱丝杠",
            "特斯拉", "Optimus", "擎天柱",
            "边缘推理", "边缘计算", "端侧AI", "NPU",
            "波士顿动力", "Figure AI", "优必选", "宇树科技", "智元机器人",
            "小批量量产", "BOM成本下降", "工厂实测", "物流场景",
            "端到端模型上车", "仿真训练平台",
            "电机", "关节模组", "编码器", "空心杯电机",
            "拓普集团", "三花智控", "绿的谐波", "双环传动",
            "鸣志电器", "汇川技术", "禾川科技",
            "机器人大模型", "RT-2", "通用机器人",
            "人形机器人量产", "万台产能"
        ]
    },
    "D": {
        "name": "大厂AI软件应用落地",
        "keywords": [
            "大模型", "Token", "Agent", "AI Agent", "RAG",
            "MaaS", "DAU", "MAU", "月活", "日活",
            "Capex", "资本开支", "资本支出",
            "API调用", "API收入", "Token消耗",
            "多模态", "多模态模型", "GPT-5", "Claude", "Gemini",
            "AI应用", "AI落地", "企业级AI",
            "AI PC", "端侧NPU", "AI手机", "AI芯片",
            "Kimi", "文心一言", "通义千问", "智谱", "百川",
            "豆包", "DeepSeek", "Qwen", "GLM",
            "AIGC", "生成式AI", "Copilot", "AI编程",
            "付费转化", "正向毛利", "商业化",
            "微软", "谷歌", "Meta", "OpenAI", "Anthropic",
            "BAT", "字节跳动", "华为盘古",
            "算力基建", "智算中心", "东数西算",
            "IDC", "数据中心",
            "国家统计局", "ICT投资", "数字经济"
        ]
    }
}

# 通用AI关键词（不特定于某条主线）
GENERAL_AI_KEYWORDS = [
    "AI", "人工智能", "算力", "深度学习", "神经网络", "机器学习",
    "大模型", "LLM", "AIGC", "生成式AI", "智能"
]

# 感兴趣但可能不符合的领域（补充新闻用）
INTERESTING_KEYWORDS = [
    "自动驾驶", "智能驾驶", "FSD", "Robotaxi",
    "量子计算", "量子芯片",
    "低空经济", "无人机", "eVTOL",
    "苹果", "Vision Pro", "AR", "VR", "MR",
    "Sora", "视频生成",
    "AI制药", "AI医疗", "AI药物发现",
    "数字人", "虚拟人",
    "6G", "卫星互联网", "星链",
    "固态电池", "钠离子电池",
    "碳化硅", "SiC", "第三代半导体",
    "RISC-V", "开源芯片",
    "商业航天", "民营火箭", "SpaceX",
    "脑机接口", "Neuralink"
]

# ========== 采集函数 ==========

def parse_date(entry):
    """解析RSS条目的日期"""
    for attr in ["published_parsed", "updated_parsed"]:
        val = getattr(entry, attr, None)
        if val:
            try:
                return datetime(*val[:6])
            except:
                pass
    return datetime.now()

def fetch_rss(name, rss_url):
    """从RSS源获取新闻（先用requests下载XML，避免SSL问题）"""
    entries = []
    print(f"  [RSS] {name}: {rss_url}")
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                          "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        }
        resp = requests.get(rss_url, headers=headers, timeout=15, verify=False)
        resp.encoding = "utf-8"
        if resp.status_code != 200:
            print(f"    ❌ HTTP {resp.status_code}")
            return []
        feed = feedparser.parse(resp.text)
        if not feed.entries:
            print(f"    ⚠️ 无条目")
            return []
        for entry in feed.entries:
            title = entry.get("title", "").strip()
            link = entry.get("link", "").strip()
            summary = entry.get("summary", entry.get("description", "")).strip()
            summary = re.sub(r'<[^>]+>', '', summary)
            pub_date = parse_date(entry)
            entries.append({
                "source": name,
                "title": title,
                "summary": summary[:500],
                "link": link,
                "pub_date": pub_date
            })
        print(f"    ✅ 获取 {len(entries)} 条")
    except requests.exceptions.SSLError:
        print(f"    ⚠️ SSL错误，尝试无验证...")
        try:
            resp = requests.get(rss_url, headers=headers, timeout=15, verify=False)
            resp.encoding = "utf-8"
            feed = feedparser.parse(resp.text)
            for entry in feed.entries:
                title = entry.get("title", "").strip()
                link = entry.get("link", "").strip()
                summary = entry.get("summary", entry.get("description", "")).strip()
                summary = re.sub(r'<[^>]+>', '', summary)
                pub_date = parse_date(entry)
                entries.append({
                    "source": name, "title": title, "summary": summary[:500],
                    "link": link, "pub_date": pub_date
                })
            print(f"    ✅ 获取 {len(entries)} 条")
        except Exception as e2:
            print(f"    ❌ 仍失败: {e2}")
    except Exception as e:
        print(f"    ⚠️ 错误: {e}")
    return entries

def fetch_eastmoney():
    """从东方财富获取全球资讯"""
    entries = []
    print(f"  [API] 东方财富全球资讯...")
    try:
        for page in range(1, 4):
            result = eastmoney_news(page=page, page_size=20)
            if result is None:
                print(f"    ⚠️ 第{page}页返回None，跳过")
                continue
            if not result:
                break
            for item in result:
                if not item:
                    continue
                title = item.get("title", item.get("art_title", "")).strip()
                if not title:
                    continue
                summary = item.get("summary", item.get("art_stitle", "")).strip()
                summary = re.sub(r'<[^>]+>', '', summary)
                pub_time_str = item.get("showTime", item.get("art_time", ""))
                pub_date = datetime.now()
                if pub_time_str:
                    try:
                        pub_date = datetime.strptime(str(pub_time_str)[:10], "%Y-%m-%d")
                    except:
                        pass
                entries.append({
                    "source": "东方财富资讯",
                    "title": title,
                    "summary": summary[:500],
                    "link": item.get("url", item.get("art_url", "")),
                    "pub_date": pub_date
                })
            time.sleep(1.5)
        print(f"    ✅ 获取 {len(entries)} 条")
    except Exception as e:
        print(f"    ⚠️ 东财错误: {e}")
    return entries

def fetch_all():
    """采集所有信源"""
    all_entries = []

    # 1. RSS信源
    for name, config in RSS_SOURCES.items():
        entries = fetch_rss(name, config["url"])
        all_entries.extend(entries)
        time.sleep(random.uniform(0.5, 1.5))

        # 如果主URL失败，尝试备用URL
        if len(entries) == 0 and config.get("alt_urls"):
            print(f"  [RSS] {name} 尝试备用URL...")
            for alt_url in config["alt_urls"]:
                entries = fetch_rss(name, alt_url)
                all_entries.extend(entries)
                if entries:
                    break
                time.sleep(random.uniform(0.5, 1.5))

    # 2. 东方财富API
    entries = fetch_eastmoney()
    all_entries.extend(entries)

    # 去重（按标题）
    seen_titles = set()
    deduped = []
    for entry in all_entries:
        title_key = entry["title"][:50].strip().lower()
        if title_key and title_key not in seen_titles:
            seen_titles.add(title_key)
            deduped.append(entry)

    print(f"\n📊 总计采集 {len(all_entries)} 条，去重后 {len(deduped)} 条")
    return deduped

# ========== 分类与筛选 ==========

def match_mainline(title, summary):
    """判断信息归属哪条主线，返回 [(主线, 细分环节, 关键词)]"""
    text = (title + " " + summary).lower()
    results = []
    for ml_id, config in KEYWORDS_MAINLINE.items():
        for kw in config["keywords"]:
            if kw.lower() in text:
                results.append((ml_id, config["name"], kw))
    if not results:
        # 通用AI关键词
        for kw in GENERAL_AI_KEYWORDS:
            if kw.lower() in text:
                results.append(("D", "大厂AI软件应用落地", kw))
                break
    return results

def is_general_ai_related(title, summary):
    text = (title + " " + summary).lower()
    for kw in GENERAL_AI_KEYWORDS:
        if kw.lower() in text:
            return True
    return False

def is_interesting_extra(title, summary):
    text = (title + " " + summary).lower()
    for kw in INTERESTING_KEYWORDS:
        if kw.lower() in text:
            return True
    return False

# ========== Excel生成 ==========

def create_excel(primary_news, secondary_news):
    """生成带打标标签的Excel（使用xlsxwriter引擎）"""
    import xlsxwriter

    all_items = [(item, "主集100") for item in primary_news] + \
                [(item, "补充50") for item in secondary_news]
    cutoff = datetime.now() - timedelta(days=FETCH_DAYS)

    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ai_news_for_labeling.xlsx")
    workbook = xlsxwriter.Workbook(output_path)

    # === Sheet 1: 打标数据 ===
    ws1 = workbook.add_worksheet("AI算力产业链新闻(打标)")

    header_fmt = workbook.add_format({
        'bold': True, 'font_color': 'white', 'bg_color': '#1F4E79',
        'align': 'center', 'valign': 'vcenter', 'border': 1, 'text_wrap': True
    })
    cell_fmt = workbook.add_format({'border': 1, 'valign': 'top', 'text_wrap': True})
    cell_fmt_yellow = workbook.add_format({
        'border': 1, 'valign': 'top', 'text_wrap': True, 'bg_color': '#FFF2CC'
    })

    headers = [
        "序号", "信源", "发布时间", "标题", "摘要",
        "系统预判主线", "系统预判细分环节",
        "你的判断-主线归属", "你的判断-触发条件",
        "你的判断-价值评分(1-5)", "你的判断-置信度",
        "链接", "类型"
    ]

    col_widths = [5, 14, 12, 50, 60, 18, 22, 22, 18, 16, 14, 50, 10]
    for i, w in enumerate(col_widths):
        ws1.set_column(i, i, w)

    for i, h in enumerate(headers):
        ws1.write(0, i, h, header_fmt)

    ws1.freeze_panes(1, 0)

    row_idx = 1
    for idx, (item, ntype) in enumerate(all_items, 1):
        title = item["title"]
        summary = item.get("summary", "")
        pub_date = item.get("pub_date", datetime.now())

        if isinstance(pub_date, datetime) and pub_date < cutoff and ntype == "主集100":
            continue

        matches = match_mainline(title, summary)
        if matches:
            mainlines = ", ".join(set(m[0] for m in matches))
            segments = ", ".join(set(m[2] for m in matches))
        elif ntype == "补充50":
            mainlines = "可能感兴趣"
            segments = "非四条主线"
        else:
            mainlines = "未识别"
            segments = ""

        date_str = pub_date.strftime("%Y-%m-%d") if isinstance(pub_date, datetime) else str(pub_date)

        values = [
            idx, item["source"], date_str, title, summary[:300],
            mainlines, segments,
            "", "", "", "",
            str(item.get("link", "")), ntype
        ]

        fmt = cell_fmt_yellow if ntype == "补充50" else cell_fmt
        for col, val in enumerate(values):
            ws1.write(row_idx, col, val, fmt)

        row_idx += 1

        if idx % 50 == 0:
            print(f"  已写入 {idx} 条...")

    # === Sheet 2: 打标说明 ===
    ws2 = workbook.add_worksheet("打标说明")
    ws2.set_column(0, 0, 80)
    instr_fmt = workbook.add_format({'text_wrap': True})
    bold_fmt = workbook.add_format({'bold': True, 'font_size': 13, 'font_color': '#1F4E79', 'text_wrap': True})

    instructions = [
        [bold_fmt, "📋 人工打标标签说明 - 请逐条标记以下4个字段"],
        [instr_fmt, ""],
        [instr_fmt, "=== 一、主线归属判断（第8列）==="],
        [instr_fmt, "填写：A / B / C / D / 跨主线 / 不相关"],
        [instr_fmt, "  A = 国产替代自主可控"],
        [instr_fmt, "  B = 英伟达全球供应链"],
        [instr_fmt, "  C = 具身智能产业化"],
        [instr_fmt, "  D = 大厂AI软件应用落地"],
        [instr_fmt, "  跨主线 = 同时影响2条以上"],
        [instr_fmt, "  不相关 = 与四条主线无关"],
        [instr_fmt, ""],
        [instr_fmt, "=== 二、触发条件编号（第9列）==="],
        [instr_fmt, "填写：1 / 2 / 3 / 4 / 5 / 0（未达触发条件）"],
        [instr_fmt, "  1=主线A：国产替代头部客户验证/批量出货/良率达标/获国家级资金"],
        [instr_fmt, "  2=主线B：进入英伟达下一代供应链/份额提升/独家匹配"],
        [instr_fmt, "  3=主线C：具身智能成本降30%+/小批量量产/场景运行超1000h"],
        [instr_fmt, "  4=主线D：Token消耗周环比增>50%/Agent正向毛利/Capex上调>20%"],
        [instr_fmt, "  5=跨主线共振：同一事件影响2条以上主线"],
        [instr_fmt, "  0=无/未达触发条件"],
        [instr_fmt, ""],
        [instr_fmt, "=== 三、价值评分（第10列）==="],
        [instr_fmt, "填写：1-5分"],
        [instr_fmt, "  1=完全无关噪声    2=边缘相关无实质影响"],
        [instr_fmt, "  3=有参考价值      4=重要结构性信号"],
        [instr_fmt, "  5=里程碑事件，可能改变投资逻辑"],
        [instr_fmt, ""],
        [instr_fmt, "=== 四、信息置信度（第11列）==="],
        [instr_fmt, "填写：高 / 中 / 低"],
        [instr_fmt, "  高=官方公告/政府文件/上市公司公告"],
        [instr_fmt, "  中=行业媒体深度报道/分析师研究/产业链调研"],
        [instr_fmt, "  低=传言/社交媒体/未经证实"],
        [instr_fmt, ""],
        [instr_fmt, "📌 示例：'华为昇腾910C通过阿里云验证'"],
        [instr_fmt, "  主线=A  触发条件=1  评分=4  置信度=高"],
        [instr_fmt, ""],
        [instr_fmt, "📌 示例：'英伟达GB200确认量产'"],
        [instr_fmt, "  主线=B  触发条件=2  评分=4  置信度=高"],
        [instr_fmt, ""],
        [instr_fmt, "⏱ 预计耗时：每条约15-30秒，150条约30-60分钟"],
    ]

    for i, (fmt, text) in enumerate(instructions):
        ws2.write(i, 0, text, fmt)

    workbook.close()
    print(f"\n✅ Excel已保存: {output_path}")
    return output_path


# ========== 主流程 ==========

def main():
    print("=" * 60)
    print("  AI算力产业链新闻采集 — 150条打标数据集")
    print("  信源：15个白名单RSS + 东方财富API")
    print(f"  关键词：{sum(len(v['keywords']) for v in KEYWORDS_MAINLINE.values())}个")
    print(f"  回溯：{FETCH_DAYS}天")
    print("=" * 60)

    # Step 1: 采集
    print("\n📡 Step 1: 采集新闻...")
    all_entries = fetch_all()

    # Step 2: 分类
    print("\n📊 Step 2: 分类排序...")
    primary = []  # 符合要求的
    secondary = []  # 补充的
    remaining_secondary = []

    for entry in all_entries:
        matches = match_mainline(entry["title"], entry.get("summary", ""))
        if matches:
            primary.append(entry)
        elif is_general_ai_related(entry["title"], entry.get("summary", "")):
            primary.append(entry)
        elif is_interesting_extra(entry["title"], entry.get("summary", "")):
            remaining_secondary.append(entry)
        else:
            remaining_secondary.append(entry)

    # 从remaining_secondary补充到secondary
    # 再去掉primary中与remaining_secondary重复的标题
    primary_titles = set(e["title"][:40].lower() for e in primary)

    # 即使不完全符合AI，也取一些作为补充
    for entry in all_entries:
        if entry not in primary:
            key = entry["title"][:40].lower()
            if key not in primary_titles:
                secondary.append(entry)
                primary_titles.add(key)

    # 主集取前100条
    primary_sorted = sorted(primary, key=lambda x: x.get("pub_date", datetime.now()), reverse=True)
    primary_100 = primary_sorted[:100]

    # 补充取50条（从remaining_secondary中取）
    # 优先取有interesting keywords的
    interesting_first = sorted(
        [e for e in secondary if is_interesting_extra(e["title"], e.get("summary", ""))],
        key=lambda x: x.get("pub_date", datetime.now()), reverse=True
    )
    other_secondary = sorted(
        [e for e in secondary if not is_interesting_extra(e["title"], e.get("summary", ""))],
        key=lambda x: x.get("pub_date", datetime.now()), reverse=True
    )
    secondary_50 = (interesting_first + other_secondary)[:50]

    print(f"\n  主集（符合要求）：{len(primary_100)} 条")
    print(f"  补充（可能感兴趣）：{len(secondary_50)} 条")

    # 统计主集各主线分布
    ml_count = {"A": 0, "B": 0, "C": 0, "D": 0}
    for entry in primary_100:
        matches = match_mainline(entry["title"], entry.get("summary", ""))
        for m in matches:
            ml_count[m[0]] = ml_count.get(m[0], 0) + 1
    print(f"  主集分布: {ml_count}")

    # Step 3: 生成Excel
    print("\n📝 Step 3: 生成Excel...")
    output = create_excel(primary_100, secondary_50)

    print("\n" + "=" * 60)
    print("  ✅ 完成！")
    print(f"  📁 {output}")
    print(f"  主集100条 + 补充50条 = 150条")
    print("  📊 第二个Sheet「打标说明」已包含完整标记指引")
    print("=" * 60)


if __name__ == "__main__":
    main()
