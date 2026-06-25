"""验证已生成的Excel文件"""
import openpyxl, os

path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ai_news_for_labeling.xlsx")
print(f"📂 文件: {path}")
print(f"📏 大小: {os.path.getsize(path)} bytes")

wb = openpyxl.load_workbook(path)
ws = wb.active
print(f"\n✅ 打开成功!")
print(f"📊 Sheet: {ws.title}")
print(f"📊 行数: {ws.max_row}, 列数: {ws.max_column}")

print("\n=== 表头 ===")
for c in range(1, ws.max_column + 1):
    print(f"  Col{c}: {ws.cell(1, c).value}")

print("\n=== 前5行样本 ===")
for r in range(2, min(7, ws.max_row + 1)):
    vals = [str(ws.cell(r, c).value or "")[:40] for c in [1, 2, 3, 4, 6, 13]]
    print(f"  #{vals[0]} | {vals[1]} | {vals[2]} | {vals[3]}")

print("\n=== 按类型统计 ===")
tc = {}
for r in range(2, ws.max_row + 1):
    t = ws.cell(r, 13).value or "未知"
    tc[t] = tc.get(t, 0) + 1
for t, c in sorted(tc.items()):
    print(f"  {t}: {c}")

print("\n=== 按信源统计 ===")
sc = {}
for r in range(2, ws.max_row + 1):
    s = ws.cell(r, 2).value or "未知"
    sc[s] = sc.get(s, 0) + 1
for s, c in sorted(sc.items(), key=lambda x: -x[1]):
    print(f"  {s}: {c}")

print("\n✅ 验证完成!")
